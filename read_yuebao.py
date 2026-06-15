#!/usr/bin/env python3
"""Extract data from monthly report files."""
import sys
from pathlib import Path

base = Path("/Users/zhengbo/Desktop/月报/1A 2026")

# 1. First check the 月报制作.xlsx in parent dir
from openpyxl import load_workbook
wb = load_workbook("/Users/zhengbo/Desktop/月报/月报制作.xlsx")
print("=== 月报制作.xlsx ===")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\nSheet: {name} - rows={ws.max_row} cols={ws.max_column}")
    for row in ws.iter_rows(max_row=30, values_only=True):
        vals = [str(v)[:60] if v is not None else "" for v in row]
        if any(v for v in vals):
            print(" | ".join(vals))

print("\n\n=== 员工个人月报文档 ===")

# 2. Read individual employee docx files
from docx import Document

doc_files = [
    "陈嘉阳4月报.docx",
    "陈嘉阳5月月报.docx",
    "吴超4月月报.docx",
    "吴超5月份月报.docx",
    "晏军5月月报.docx",
    "张建强4月 工作月报.docx",
    "张建强5月 工作月报.docx",
]

for fname in doc_files:
    path = base / fname
    if not path.exists():
        print(f"\n--- {fname}: NOT FOUND ---")
        continue
    try:
        doc = Document(str(path))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        print(f"\n--- {fname} ({len(text)} chars) ---")
        print(text[:2000])
        if len(text) > 2000:
            print(f"... [truncated, {len(text)-2000} more chars]")
    except Exception as e:
        print(f"\n--- {fname}: ERROR - {e} ---")

# Also read the .txt file for 晏军4月
print("\n\n=== 晏军4月（月报）.txt ===")
try:
    txt = (base / "晏军4月（月报）.txt").read_text(encoding="utf-8")
    print(txt)
except Exception as e:
    print(f"Error: {e}")
